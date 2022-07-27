import glob
import os
import datetime
import math
from django.core.files import File

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from bibstat import settings

from data.principals import principal_for_library_type
from libstat.models import Survey, OpenData, Variable

import logging

DATE_FORMAT = "%Y_%m_%d_%H_%M_%S"


logger = logging.getLogger(__name__)


def _cache_dir_path():
    if settings.ENVIRONMENT == "local":
        return "{}/data/excel_exports/".format(os.getcwd())
    return "/data/appl/excel_exports/"


def _cache_path(year, file_name_str="public_export_{} {}.xslx", date_str=None):
    file_name = file_name_str.format(year, date_str if date_str else "*")
    if settings.ENVIRONMENT == "local":
        return "{}/data/excel_exports/{}".format(os.getcwd(), file_name)
    else:
        return "/data/appl/excel_exports/{}".format(file_name)


def _cached_workbook_exists_and_is_valid(
    year, file_name="public_export_{} {}.xslx", workbook_is_public=True
):
    paths = sorted(glob.glob(_cache_path(year, file_name_str=file_name)))
    if not paths:
        return False

    cache_date = datetime.datetime.strptime(
        paths[-1].split(" ")[-1].split(".")[0], DATE_FORMAT
    ).replace(tzinfo=datetime.timezone.utc)

    if workbook_is_public:
        latest_modification = OpenData.objects.first().date_modified
    else:
        latest_modification = (
            Survey.objects.all().order_by("-date_modified").first().date_modified
        )

    return cache_date > latest_modification


def _cache_workbook(
    workbook, year, file_name_str="public_export_{} {}.xslx", workbook_is_public=True
):
    for filename in os.listdir(_cache_dir_path()):
        if ".xslx" in filename:
            if (workbook_is_public == True and filename.startswith("public")) or (
                workbook_is_public == False and filename.startswith("survey")
            ):
                os.remove("%s%s" % (_cache_dir_path(), filename))
    with open(
        _cache_path(
            year, file_name_str, datetime.datetime.utcnow().strftime(DATE_FORMAT)
        ),
        "wb",
    ) as f:
        File(f).write(save_virtual_workbook(workbook))


def public_excel_workbook(year):
    if not _cached_workbook_exists_and_is_valid(year):
        _cache_workbook(_published_open_data_as_workbook(year), year)
    return sorted(glob.glob(_cache_path(year)))[-1]


def _published_open_data_as_workbook(year):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Värden"

    public_variables = list(Variable.objects.filter(is_public=True).distinct("key"))

    variable_keys = list(
        OpenData.objects.filter(
            is_active=True, sample_year=year, variable_key__in=public_variables
        ).distinct("variable_key")
    )
    sigels = list(
        OpenData.objects.filter(
            is_active=True, sample_year=year, variable_key__in=public_variables
        ).distinct("sigel")
    )

    libraries = {}
    for sigel in sigels:
        libraries[sigel] = dict.fromkeys(variable_keys)

    for open_data in OpenData.objects.filter(
        is_active=True, sample_year=year, variable_key__in=public_variables
    ).only("library_name", "variable_key", "sigel", "value"):
        libraries[open_data.sigel][open_data.variable_key] = open_data.value

    header = ["Bibliotek", "Sigel", "Bibliotekstyp", "Kommunkod", "Stad", "Externt id"]
    variable_index = {}
    for index, key in enumerate(variable_keys):
        variable_index[key] = index + 6
        header.append(key)
    worksheet.append(header)

    for sigel in libraries:
        # library = Survey.objects.filter(library__sigel=sigel, sample_year=year).first().library
        library = Survey.objects.filter(library__sigel=sigel, sample_year=year).first()
        if library != None:
            library = library.library
        else:
            continue  # KP 180419
        row = [""] * len(header)
        row[0] = library.name
        row[1] = (
            sigel if year >= 2014 else ""
        )  # Do not show auto-generated sigels (used before 2014)
        row[2] = library.library_type
        row[3] = library.municipality_code
        row[4] = library.city
        row[5] = (
            library.external_identifiers[0].identifier
            if library.external_identifiers
            and len(library.external_identifiers) > 0
            and library.external_identifiers[0].identifier
            else ""
        )

        for key in variable_keys:
            row[variable_index[key]] = libraries[sigel][key]
        worksheet.append(row)

    variable_sheet = workbook.create_sheet()
    variable_sheet.title = "Definitioner"
    for variable in Variable.objects.filter(key__in=variable_keys):
        variable_sheet.append([variable.key, variable.description])

    return workbook


def _populate_survey_cells(survey, worksheet, headers_columns_dict, row_no):
    if not survey:
        return
    worksheet.cell(
        row=row_no, column=headers_columns_dict["År"]
    ).value = survey.sample_year
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Bibliotek"]
    ).value = survey.library.name
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Sigel"]
    ).value = survey.library.sigel
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Bibliotekstyp"]
    ).value = survey.library.library_type
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Status"]
    ).value = Survey.status_label(survey.status)
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Email"]
    ).value = survey.library.email
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Kommunkod"]
    ).value = survey.library.municipality_code
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Stad"]
    ).value = survey.library.city
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Adress"]
    ).value = survey.library.address
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Postkod"]
    ).value = survey.library.zip_code
    worksheet.cell(row=row_no, column=headers_columns_dict["Huvudman"]).value = (
        principal_for_library_type[survey.library.library_type]
        if survey.library.library_type in principal_for_library_type
        else None
    )
    worksheet.cell(row=row_no, column=headers_columns_dict["Kan publiceras?"]).value = (
        "Ja"
        if survey.can_publish()
        else "Nej: " + survey.reasons_for_not_able_to_publish()
    )
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Samredovisar andra bibliotek"]
    ).value = ("Ja" if survey.is_reporting_for_others() else "Nej")
    worksheet.cell(row=row_no, column=headers_columns_dict["Samredovisas"]).value = (
        "Ja" if survey.is_reported_by_other() else "Nej"
    )
    worksheet.cell(
        row=row_no, column=headers_columns_dict["Redovisas av"]
    ).value = ",".join(survey.reported_by())

    for observation in survey.observations:
        variable_key = observation.variable.key
        if headers_columns_dict.get(variable_key, None):
            value = observation.value
            if observation.value_unknown:
                value = "okänt värde"
            worksheet.cell(
                row=row_no, column=headers_columns_dict[variable_key]
            ).value = value

    other_sigels = [s for s in survey.selected_libraries if s != survey.library.sigel]
    if len(other_sigels) > 0:
        other_surveys = Survey.objects.filter(
            library__sigel__in=other_sigels, sample_year=survey.sample_year
        ).only("library")
        for other_survey in other_surveys:
            worksheet.cell(
                row=row_no, column=headers_columns_dict["Samredovisat bibliotek"]
            ).value = "%s (%s)" % (
                other_survey.library.name,
                other_survey.library.sigel,
            )
            worksheet.cell(
                row=row_no, column=headers_columns_dict["Gatuadress"]
            ).value = other_survey.library.address
            worksheet.cell(
                row=row_no, column=headers_columns_dict["Postnummer"]
            ).value = other_survey.library.zip_code


def _load_surveys_and_append_worksheet_rows(
    surveys, worksheet, headers_columns_dict, offset=0, include_previous_year=False
):
    row_no = 2 + offset
    for survey in surveys:
        _populate_survey_cells(survey, worksheet, headers_columns_dict, row_no)
        if include_previous_year:
            previous = survey.previous_years_survey()
            if previous:
                _populate_survey_cells(
                    previous,
                    worksheet,
                    headers_columns_dict,
                    row_no + 1,
                )
            row_no += 2
        else:
            row_no += 1


def surveys_to_excel_workbook(survey_ids, include_previous_year=False):
    bulk_size = 300
    bulks = int(math.ceil(len(survey_ids) / bulk_size))

    headers_dict = {}

    headers = [
        "År",
        "Bibliotek",
        "Sigel",
        "Bibliotekstyp",
        "Status",
        "Email",
        "Kommunkod",
        "Stad",
        "Adress",
        "Postkod",
        "Huvudman",
        "Kan publiceras?",
        "Samredovisar andra bibliotek",
        "Samredovisas",
        "Redovisas av",
    ]
    headers += [
        str(observation.variable.key)
        for observation in Survey.objects.get(pk=survey_ids[0]).observations
        if observation.variable.key
    ]
    headers += ["Samredovisat bibliotek", "Gatuadress", "Postnummer"]

    # Create a headers dictionary for lookup of column number
    column_no = 1
    for header in headers:
        headers_dict[header] = column_no
        column_no = column_no + 1

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)

    offset = 0
    for index in range(0, bulks):
        start_index = index * bulk_size
        stop_index = start_index + bulk_size
        if stop_index > len(survey_ids):
            stop_index = len(survey_ids)

        ids = survey_ids[start_index:stop_index]

        surveys = Survey.objects.filter(id__in=ids).order_by("library__name")

        _load_surveys_and_append_worksheet_rows(
            surveys,
            worksheet,
            headers_dict,
            offset=offset,
            include_previous_year=include_previous_year,
        )
        offset += bulk_size

    return workbook
